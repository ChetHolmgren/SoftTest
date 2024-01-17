# -*- coding: utf-8 -*-
# Time:2023/12/18 11:50
# Author:张柘
# File:data.py
# Desc:

from config.config import excel_path


class ReadWrite:
    def __init__(self):
        self.txtpath = r''
        self.excelpath = excel_path
        self.yamlpath = r''

    # txt文本文件的读写
    def txtread(self):
        list1 = []
        f = open(self.txtpath, 'r', encoding='utf-8')
        values = f.readlines()
        f.close()
        for data in values:
            data_v = data.strip('\n')
            list1.append(data_v)
        return list1

    def txtwrite(self):
        f = open(self.txtpath, 'a', encoding='utf-8')
        username = input('请输入用户名：')
        password = input('请输入密码：')
        values = f"{username},{password}\n"
        f.writelines(values)
        f.close()
        return True

    # excel文件的读写
    def excelread(self, sheetname):
        import openpyxl
        wb = openpyxl.load_workbook(self.excelpath)
        table = wb[sheetname]
        rows = table.max_row
        cols = table.max_column
        list2 = []
        for row in range(2, rows + 1):
            list1 = []
            for col in range(1, cols + 1):
                value = table.cell(row, col).value
                list1.append(value)
            list2.append(list1)
        return list2

    # 多个值的变量用*表示，用元组存储
    def excelwrite(self, *args, sheetname):
        import openpyxl
        wb = openpyxl.load_workbook(self.excelpath)
        table = wb[sheetname]
        rows = table.max_row
        cols = len(args)
        for col in range(cols):
            table.cell(rows + 1, col + 1).value = args[col]
        wb.save(self.excelpath)

    @staticmethod
    def mysqlread():
        import pymysql
        db = pymysql.connect(host='localhost', port=3306, user='root', password='123456', database='test1',
                             charset='utf8')
        cur = db.cursor()
        sql = 'select * from users'
        cur.execute(sql)
        content = cur.fetchall()
        return content

    @staticmethod
    def mysqlwrite(username, password):
        import pymysql
        db = pymysql.connect(host='localhost', port=3306, user='root', password='123456', database='test1',
                             charset='utf8')
        cur = db.cursor()
        sql = "insert into users values('%s','%s')" % (username, password)
        cur.execute(sql)
        db.commit()

    # yaml文件的读写
    # 非python自带，需要提前安装pyyaml
    # 安装命令：python -m pip install pyyaml
    # yml文件语法
    # 使用缩进表示层级
    # 缩进使用空格键，不是tab键
    # 区分大小写
    # #表示注释
    # -开头的数据会转化成列表
    # 读写操作
    # 读 打开文件 f = open(file,模式，encoding = 'utf-8'
    # content = f.read()
    # yaml.safe_load(content) 转换成字典格式
    # 写 打开文件 f = open(file,模式，encoding = 'utf-8'
    # 写操作 准备的数据：字典的格式
    # yaml.safe_dump(数据,f)
    # 关闭文件 f.close()

    def yamlread(self):
        import yaml
        with open(self.yamlpath, 'r', encoding='utf-8') as f:
            content = f.read()
            data = yaml.safe_load(content)
            f.close()
            return data

    def yamlwrite(self, username, password):
        import yaml
        with open(self.yamlpath, 'a', encoding='utf-8') as f:
            data = {
                'username': username, 'password': password
            }
            yaml.safe_dump(data, f)
            f.close()
            return True
